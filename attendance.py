from tkinter import *
from PIL import Image, ImageTk
from tkinter import ttk
from tkinter import messagebox
import mysql.connector
import cv2
import os
import csv
import openpyxl
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

mydata=[]
class Attendance:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1530x790+0+0")
        self.root.title("Face Recognition System")

        #--------variables
        self.var_atten_id=StringVar()
        self.var_atten_fun=StringVar()
        self.var_atten_name=StringVar()
        self.var_atten_dep=StringVar()
        self.var_atten_time=StringVar()
        self.var_atten_date=StringVar()
        self.var_atten_attendance=StringVar()

         # first image
        img = Image.open(r"D:\SmartAttends\face-detection\img\sys.png")
        img = img.resize((1290, 230))
        self.photoimg = ImageTk.PhotoImage(img)
        f_lbl = Label(self.root, image=self.photoimg)
        f_lbl.place(x=0, y=0, width=1290, height=230)

        #bg image
        img3 = Image.open(r"D:\SmartAttends\face-detection\img\3.png")
        img3 = img3.resize((1530, 500))
        self.photoimg3 = ImageTk.PhotoImage(img3)
        bg_img = Label(self.root, image=self.photoimg3)
        bg_img.place(x=0, y=232, width=1290, height=500)

        title_lbl=Label(self.root,text="ATTENDANCE MANAGEMENT SYSTEM",font=("times new roman",35,"bold"),bg="white",fg="green")
        title_lbl.place(x=0,y=232,width=1290,height=45)

        
        main_frame=Frame(bg_img,bd=2,bg="white")
        main_frame.place(x=5,y=50,width=1260,height=400)

         #Left label fram
        Left_frame=LabelFrame(main_frame,bd=2,relief=RIDGE,text="Employees Attendance Details",font=("times new roman",12,"bold"),fg="red")
        Left_frame.place(x=10,y=7,width=610,height=350)

        

        left_inside_frame=Frame(main_frame,bd=2,bg="white",relief=RIDGE)
        left_inside_frame.place(x=20,y=39,width=590,height=310)

        #labelLandEntry
        #attendance id
        studentId_label = Label(left_inside_frame, text="Attendance ID:", font=("times new roman", 13, "bold"),bg="white")
        studentId_label.grid(row=0,column=0,padx=10,pady=7,sticky=W)

        studntID_entry=ttk.Entry(left_inside_frame,textvariable=self.var_atten_id,width=17,font=("times new roman", 13, "bold"))
        studntID_entry.grid(row=0,column=1,padx=10,pady=7,sticky=W)

        #name employe
        studentName_label = Label(left_inside_frame, text="Function:", font=("times new roman", 13, "bold"),bg="white")
        studentName_label.grid(row=0,column=2,padx=10,pady=7,sticky=W)

        dep_combo=ttk.Combobox(left_inside_frame,textvariable=self.var_atten_fun,font=("times new roman",12,"bold"),width=17,state="readonly")
        dep_combo["values"]=("Select Function","finition","apprenti","ouvrier")
        dep_combo.current(0)
        dep_combo.grid(row=0,column=3,padx=10,pady=7,sticky=W)

        #sexe employe
        studentName_label = Label(left_inside_frame, text="Name:", font=("times new roman", 13, "bold"),bg="white")
        studentName_label.grid(row=1,column=0,padx=10,pady=7,sticky=W)

        studentName_entry=ttk.Entry(left_inside_frame,textvariable=self.var_atten_name,width=17,font=("times new roman", 13, "bold"))
        studentName_entry.grid(row=1,column=1,padx=10,pady=7,sticky=W)

        


        #cin employe
        studentName_label = Label(left_inside_frame, text="Service:", font=("times new roman", 13, "bold"),bg="white")
        studentName_label.grid(row=1,column=2,padx=10,pady=5,sticky=W)

        studentName_entry=ttk.Entry(left_inside_frame,textvariable=self.var_atten_dep,width=17,font=("times new roman", 13, "bold"))
        studentName_entry.grid(row=1,column=3,padx=10,pady=5,sticky=W)

         #function employe
        studentId_label = Label(left_inside_frame, text="Time:", font=("times new roman", 13, "bold"),bg="white")
        studentId_label.grid(row=2,column=0,padx=10,pady=7,sticky=W)

        studntID_entry=ttk.Entry(left_inside_frame,textvariable=self.var_atten_time,width=17,font=("times new roman", 13, "bold"))
        studntID_entry.grid(row=2,column=1,padx=10,pady=7,sticky=W)

        #statut employe
        studentName_label = Label(left_inside_frame, text="Date:", font=("times new roman", 13, "bold"),bg="white")
        studentName_label.grid(row=2,column=2,padx=10,pady=5,sticky=W)

        studentName_entry=ttk.Entry(left_inside_frame,textvariable=self.var_atten_date,width=17,font=("times new roman", 13, "bold"))
        studentName_entry.grid(row=2,column=3,padx=10,sticky=W)

         #statut employe
        studentName_label = Label(left_inside_frame, text="Att status:", font=("times new roman", 13, "bold"),bg="white")
        studentName_label.grid(row=3,column=0,padx=10,pady=5,sticky=W)


        dep_combo=ttk.Combobox(left_inside_frame,textvariable=self.var_atten_attendance,font=("times new roman",12,"bold"),width=17,state="readonly")
        dep_combo["values"]=("Status","Present","Absent")
        dep_combo.current(0)
        dep_combo.grid(row=3,column=1,padx=10,pady=7,sticky=W)
        #btn
        btn_frame=Frame(left_inside_frame,bd=2,relief=RIDGE,bg="white")
        btn_frame.place(x=0,y=200,width=596,height=35)

        #save btn
        save_btn=Button(btn_frame,text="Import xlsx",command=self.importExcel,width=14,font=("times new roman", 13, "bold"),bg="blue",fg="white")
        save_btn.grid(row=0,column=0)

        update_btn=Button(btn_frame,text="Export xlsx",command=self.exportXlsx,width=14,font=("times new roman", 13, "bold"),bg="blue",fg="white")
        update_btn.grid(row=0,column=1)

        delete_btn=Button(btn_frame,text="Update",width=14,font=("times new roman", 13, "bold"),bg="blue",fg="white")
        delete_btn.grid(row=0,column=2)

        reset_btn=Button(btn_frame,text="Reset",width=14,command=self.reset_data,font=("times new roman", 13, "bold"),bg="blue",fg="white")
        reset_btn.grid(row=0,column=3)

        #Right label fram
        Right_frame=LabelFrame(main_frame,bd=2,relief=RIDGE,text="Attendance Details",font=("times new roman",12,"bold"),fg="red")
        Right_frame.place(x=630,y=7,width=614,height=350)

        # left_inside_frame=Frame(main_frame,bd=2,bg="white",relief=RIDGE)
        # left_inside_frame.place(x=640,y=39,width=590,height=310)

        #create table
        table_frame = LabelFrame(Right_frame, bd=2, bg="white", relief=RIDGE)
        table_frame.place(x=5, y=10, width=600, height=310)

        #scroll bar table
        scroll_x=ttk.Scrollbar(table_frame,orient=HORIZONTAL)
        scroll_y=ttk.Scrollbar(table_frame,orient=VERTICAL)

        # Treeview for the table
        self.AttendanceReportTable = ttk.Treeview(table_frame, columns=("ID", "Function", "Name", "Service", "Time", "Date", "Attendance"),
                                                   xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)

        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)

        scroll_x.config(command=self.AttendanceReportTable.xview)
        scroll_y.config(command=self.AttendanceReportTable.yview)

        # Set column headings
        self.AttendanceReportTable.heading("ID", text="Attendance ID")
        self.AttendanceReportTable.heading("Function", text="Function")
        self.AttendanceReportTable.heading("Name", text="Name")
        self.AttendanceReportTable.heading("Service", text="Service")
        self.AttendanceReportTable.heading("Time", text="Time")
        self.AttendanceReportTable.heading("Date", text="Date")
        self.AttendanceReportTable.heading("Attendance", text="Attendance")

        self.AttendanceReportTable["show"]="headings"
        self.AttendanceReportTable.column("ID",width=100)
        self.AttendanceReportTable.column("Function",width=100)
        self.AttendanceReportTable.column("Name",width=100)
        self.AttendanceReportTable.column("Service",width=100)
        self.AttendanceReportTable.column("Time",width=100)
        self.AttendanceReportTable.column("Date",width=100)
        self.AttendanceReportTable.column("Attendance",width=100)

        self.AttendanceReportTable.pack(fill=BOTH, expand=1)

        self.AttendanceReportTable.bind("<ButtonRelease>",self.get_cursor)
                

#-------------------fetch data    
    def fetchdata(self,rows):
        self.AttendanceReportTable.delete(*self.AttendanceReportTable.get_children())
        for i in rows:
            self.AttendanceReportTable.insert("",END,values=i)
    
    #import xlsx
    def importExcel(self):
        global mydata
        mydata.clear()
        fin = filedialog.askopenfilename(initialdir=os.getcwd(), title="Open Excel", filetypes=(("Excel File", "*.xlsx"), ("All Files", "*.*")), parent=self.root)
        
        # Clear existing data before reading new data
        mydata = []
        
        workbook = openpyxl.load_workbook(fin)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            mydata.append(row)

        self.fetchdata(mydata)
    
    #export xlsx
    def exportXlsx(self):
        try:
            global mydata

            if len(mydata) < 1:
                messagebox.showerror("No data", "No Data found to export", parent=self.root)
                return False

            fln = filedialog.asksaveasfilename(initialdir=os.getcwd(), title="Save Excel", filetypes=(("Excel File", "*.xlsx"), ("All Files", "*.*")), parent=self.root)

            workbook = Workbook()
            sheet = workbook.active

            for row_data in mydata:
                sheet.append(row_data)

            # Save the workbook to the specified file
            workbook.save(fln)

            messagebox.showinfo("Data Export", f"Your data exported to {os.path.basename(fln)} successfully")

        except Exception as es:
            messagebox.showerror("Error", f"Due To: {str(es)}", parent=self.root)

    def get_cursor(self,event=""):
        cursor_row=self.AttendanceReportTable.focus()
        content=self.AttendanceReportTable.item(cursor_row)
        rows=content['values']
        self.var_atten_id.set(rows[0])
        self.var_atten_fun.set(rows[1])
        self.var_atten_name.set(rows[2])
        self.var_atten_dep.set(rows[3])
        self.var_atten_time.set(rows[4])
        self.var_atten_date.set(rows[5])
        self.var_atten_attendance.set(rows[6])

    def reset_data(self):
        self.var_atten_id.set("")
        self.var_atten_fun.set("")
        self.var_atten_name.set("")
        self.var_atten_dep.set("")
        self.var_atten_time.set("")
        self.var_atten_date.set("")
        self.var_atten_attendance.set("")
        
if __name__ == "__main__":
    root = Tk()
    obj = Attendance(root)
    root.mainloop()
